import os
import fitz  # PyMuPDF
import pandas as pd
from openpyxl import Workbook

def extract_text_based_table(page):
    words = page.get_text("words")  # Extract words with coordinates
    words.sort(key=lambda w: (w[1], w[0]))  # Sort by y-coordinates (top to bottom), then x-coordinates (left to right)
    
    rows = []
    current_row = []
    last_y = None
    
    for w in words:
        x0, y0, x1, y1, word, *_ = w
        word = word.strip()
        if not word:
            continue
        if last_y is None:
            last_y = y0
        if abs(y0 - last_y) > 5:  # New row detection based on vertical gap
            if current_row:
                rows.append(current_row)
            current_row = [word]
            last_y = y0
        else:
            current_row.append(word)
    
    if current_row:
        rows.append(current_row)
    
    return rows

def extract_tables_from_pdf(pdf_path, output_excel_path):
    doc = fitz.open(pdf_path)
    workbook = Workbook()
    
    for page_num, page in enumerate(doc):
        table_data = extract_text_based_table(page)
        if table_data:
            sheet = workbook.create_sheet(title=f"Page_{page_num + 1}")
            for row in table_data:
                sheet.append(row)
            print(f"✔️ Extracted table from Page {page_num + 1}")
        else:
            print(f"⚠️ No table detected on Page {page_num + 1}")
    
    # Remove default sheet
    if "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]
    
    workbook.save(output_excel_path)
    print(f"✅ Tables saved to: {output_excel_path}")

def process_all_pdfs(pdf_folder, output_folder):
    os.makedirs(output_folder, exist_ok=True)
    
    for filename in os.listdir(pdf_folder):
        if filename.lower().endswith(".pdf"):
            pdf_path = os.path.join(pdf_folder, filename)
            output_filename = os.path.splitext(filename)[0] + ".xlsx"
            output_excel_path = os.path.join(output_folder, output_filename)
            
            print(f"\n📄 Processing: {filename}")
            extract_tables_from_pdf(pdf_path, output_excel_path)

# Example Usage
pdf_folder = "pdfs"  # Folder containing PDFs
output_folder = "excel_output"  # Output folder for extracted tables
process_all_pdfs(pdf_folder, output_folder)
